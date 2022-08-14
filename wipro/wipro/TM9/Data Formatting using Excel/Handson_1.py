#!/usr/bin/env python
# coding: utf-8

# In[ ]:


#-->writing data in excel
import openpyxl 

from openpyxl import Workbook 
b1=Workbook()

b1['Sheet'].title='Students_data'
sh1=b1.active
sh1['A1'].value='Roll Number'
sh1['B1'].value='Fistname'
sh1['C1'].value='Lastname'
sh1['D1'].value='Mail_ID'
sh1['A2'].value='235'
sh1['B2'].value='muzim'
sh1['C2'].value='Shaik'
sh1['D2'].value='mz@gmail.com'
sh1['A3'].value='215'
sh1['B3'].value='selena'
sh1['C3'].value='Gomez'
sh1['D3'].value='sel@gmail.com'
sh1['A4'].value='236'
sh1['B4'].value='shan'
sh1['C4'].value='Shake'
sh1['D4'].value='shan@gmail.com'
sh1['A5'].value='240'
sh1['B5'].value='rahim'
sh1['C5'].value='Syed'
sh1['D5'].value='rahim@gmail.com'
#save xlsx
b1.save("student.xlsx")

#-->Reading data from existing file
from openpyxl import load_workbook
b2=load_workbook("student.xlsx")
print(b2.sheetnames)
s=b2.active
for row in s.values:
    for value in row:
        print(value,'\t',end='')
    print()

